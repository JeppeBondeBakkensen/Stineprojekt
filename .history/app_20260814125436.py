from __future__ import annotations

import json
from html import escape
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook
import pydeck as pdk
import streamlit as st


ROOT = Path(__file__).parent
GEOJSON_FILE = ROOT / "BI_BTR_RELEASEAFSNIT_OD_1972479174689089199.geojson"
NAMES_FILE = ROOT / "railway_hover_names.xlsx"
CONTRACTS_FILE = next(ROOT.glob("Kopi af Oversigt*.xlsx"))
CONTRACT_SHEET = "Kopi af Oversigt over kontrakte"


@st.cache_data
def load_railway_geometry() -> dict:
    with GEOJSON_FILE.open(encoding="utf-8") as source:
        geometry = json.load(source)

    workbook = load_workbook(NAMES_FILE, read_only=True, data_only=True)
    worksheet = workbook["Rail stretches"]
    names = {
        int(stretch_id): str(name)
        for stretch_id, name in worksheet.iter_rows(min_row=2, values_only=True)
        if stretch_id is not None and name
    }
    workbook.close()

    for feature in geometry["features"]:
        properties = feature["properties"]
        properties["HOVER_NAME"] = names.get(
            int(properties["STRKAFS"]), properties["NAVN"]
        )

    return geometry


def display_value(value: object) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


@st.cache_data
def load_contracts() -> list[dict[str, str]]:
    workbook = load_workbook(CONTRACTS_FILE, read_only=True, data_only=True)
    worksheet = workbook[CONTRACT_SHEET]
    rows = worksheet.iter_rows(min_row=1, max_col=22, values_only=True)
    headers = [display_value(value) for value in next(rows)]
    contracts = []
    for row in rows:
        contract = {
            header: display_value(value)
            for header, value in zip(headers, row)
            if header and display_value(value)
        }
        if contract.get("SAP nummer") or contract.get("Kontraktens titel"):
            contracts.append(contract)
    workbook.close()
    return contracts


def normalized_text(value: object) -> str:
    text = unicodedata.normalize("NFKD", display_value(value)).casefold()
    return " ".join(text.replace("–", "-").replace("—", "-").split())


def route_codes(properties: dict) -> set[str]:
    codes = set()
    for raw_code in re.findall(r"\d+", display_value(properties.get("TIB"))):
        number = int(raw_code)
        codes.add(str(number))
        if 1000 <= number <= 1099:
            codes.add(str(number - 1000))
        if 800 <= number <= 899 and number % 10 == 0:
            codes.add(str(number // 10))
    if properties.get("BANENR"):
        codes.add(display_value(properties["BANENR"]))
    return codes


def matching_contracts(properties: dict, contracts: list[dict[str, str]]) -> list[dict[str, str]]:
    stretch_name = normalized_text(properties.get("NAVN"))
    codes = route_codes(properties)
    matches = []
    for contract in contracts:
        geography = normalized_text(
            contract.get(
                "Geografisk område + TIB-strækninger med responstid for enkelt strækning",
                "",
            )
        )
        nationwide = any(
            phrase in geography
            for phrase in ("hele landet", "landsdaekkende", "landsdækkende")
        )
        named_match = stretch_name and stretch_name in geography
        code_match = any(
            re.search(rf"(?<!\d){re.escape(code)}(?!\d)", geography)
            for code in codes
        )
        if nationwide or named_match or code_match:
            matches.append(contract)
    return matches


def tooltip_text(value: object, limit: int = 180) -> str:
    text = " ".join(display_value(value).split())
    if len(text) > limit:
        text = f"{text[: limit - 1].rstrip()}…"
    return escape(text or "—")


def add_contract_tooltips(geometry: dict, contracts: list[dict[str, str]]) -> None:
    for feature in geometry["features"]:
        properties = feature["properties"]
        matches = matching_contracts(properties, contracts)
        start_km = properties.get("FRA_KM")
        end_km = properties.get("TIL_KM")
        length_km = (
            f"{abs(float(end_km) - float(start_km)):.1f} km"
            if start_km is not None and end_km is not None
            else "—"
        )
        properties["TOOLTIP_STRETCH"] = tooltip_text(properties.get("HOVER_NAME"))
        properties["TOOLTIP_CONTEXT"] = tooltip_text(
            properties.get("DELSTRK_NAVN") or properties.get("HVDSTRK_NAVN")
        )
        properties["TOOLTIP_LENGTH"] = length_km
        properties["TOOLTIP_TYPE"] = tooltip_text(properties.get("BTRTYPETEKST"), 50)
        properties["TOOLTIP_CLASS"] = tooltip_text(
            properties.get("BANEKLASSETEKST"), 50
        )
        properties["CONTRACT_COUNT"] = len(matches)
        if not matches:
            properties.update(
                {
                    "CONTRACT_TITLE": "No matching contract",
                    "CONTRACT_SUPPLIER": "—",
                    "CONTRACT_SAP": "—",
                    "CONTRACT_RESPONSE": "—",
                    "CONTRACT_MORE": "Click the stretch to inspect its route data",
                }
            )
            continue

        primary = matches[0]
        additional = len(matches) - 1
        properties.update(
            {
                "CONTRACT_TITLE": tooltip_text(primary.get("Kontraktens titel")),
                "CONTRACT_SUPPLIER": tooltip_text(primary.get("Leverandørnavn")),
                "CONTRACT_SAP": tooltip_text(primary.get("SAP nummer"), 60),
                "CONTRACT_RESPONSE": tooltip_text(primary.get("Responstid")),
                "CONTRACT_MORE": (
                    f"+{additional} additional contract{'s' if additional != 1 else ''} · click for details"
                    if additional
                    else "Click for full contract details"
                ),
            }
        )


def render_contract_panel(properties: dict | None, contracts: list[dict[str, str]]) -> None:
    st.markdown(
        '<div class="panel-kicker">BANEDANMARK · CONTRACT MAP</div>',
        unsafe_allow_html=True,
    )
    if not properties:
        st.markdown("# Select a railway stretch")
        st.markdown(
            '<div class="empty-state">'
            '<div class="empty-icon">↖</div>'
            '<strong>Explore the railway map</strong>'
            '<span>Click a black railway stretch to view matching contracts, '
            'response times and operational requirements.</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        return

    route_name = escape(display_value(properties.get("NAVN")) or "Selected stretch")
    st.markdown(
        f'<div class="route-header"><span>SELECTED STRETCH</span>'
        f'<h1>{route_name}</h1>'
        f'<div class="route-tags"><b>TIB/BTR {escape(display_value(properties.get("TIB")) or "—")}</b>'
        f'<b>Railway {escape(display_value(properties.get("BANENR")) or "—")}</b>'
        f'<b>km {escape(display_value(properties.get("FRA_KM")) or "—")}–'
        f'{escape(display_value(properties.get("TIL_KM")) or "—")}</b></div></div>',
        unsafe_allow_html=True,
    )
    matches = matching_contracts(properties, contracts)
    if not matches:
        st.warning("No matching contract was found in the contract workbook.")
        return

    st.markdown(
        f'<div class="match-count"><span>{len(matches)}</span> matching contract'
        f'{"s" if len(matches) != 1 else ""}</div>',
        unsafe_allow_html=True,
    )
    selected_index = st.selectbox(
        "Active contract",
        range(len(matches)),
        format_func=lambda index: (
            f"{matches[index].get('Kontraktens titel', 'Contract')} · "
            f"{matches[index].get('SAP nummer', 'No SAP number')}"
        ),
        key=f"contract-{properties.get('STRKAFS', 'selected')}",
    )
    contract = matches[selected_index]

    def fact_card(label: str, value: str) -> str:
        return (
            '<div class="fact-card">'
            f'<span>{escape(label)}</span><strong>{escape(value or "—")}</strong>'
            '</div>'
        )

    st.markdown(
        '<div class="fact-grid">'
        + fact_card("Supplier", contract.get("Leverandørnavn", ""))
        + fact_card("SAP number", contract.get("SAP nummer", ""))
        + fact_card("Effective from", contract.get("Ikraft-trædelse", ""))
        + fact_card(
            "Duration",
            contract.get(
                "Varighed \ni år (inkl mobilisering + option på forlængelse)", ""
            ),
        )
        + '</div>',
        unsafe_allow_html=True,
    )

    sections = [
        (
            "Coverage & response",
            [
                "Geografisk område + TIB-strækninger med responstid for enkelt strækning",
                "Responstid",
                "Beskrivelse af beredskabet",
            ],
        ),
        (
            "Capacity & operations",
            [
                "Antal fejl, som leverandøren skal kunne håndtere på én gang.",
                "Antal arbejdshold og personer pr. hold, som leverandøren har pligt til at stille med.",
                "Beskrevet fremgangsmåde omkring fejlretning",
                "Krav til data registrering",
                "Afrapporteringsfrekvens for leverandør",
            ],
        ),
        (
            "Penalties & obligations",
            [
                "Oversigt over bodsbelagte emner",
                "Beskrivelse af og i givet fald hvordan bod/boderne maksimeres",
                "Beskrivelse af boderne",
                "Henvisning til hvor der er beskrevet i kontrakten",
                "Henvisning til hvor der er beskrevet i kontrakten2",
            ],
        ),
        (
            "Ownership & definitions",
            ["Vedligeholdelsesforvalter", "GFS", "Kontraktansvarlig (leder)", "Definitioner"],
        ),
    ]
    shown_fields = {
        "Kontraktens titel",
        "SAP nummer",
        "Leverandørnavn",
        "Ikraft-trædelse",
        "Varighed \ni år (inkl mobilisering + option på forlængelse)",
    }
    for title, fields in sections:
        populated = [(field, contract.get(field, "")) for field in fields if contract.get(field)]
        shown_fields.update(fields)
        if not populated:
            continue
        with st.expander(title, expanded=title == "Coverage & response"):
            for label, value in populated:
                st.markdown(
                    f'<div class="detail-item"><span>{escape(label)}</span>'
                    f'<p>{escape(value).replace(chr(10), "<br>")}</p></div>',
                    unsafe_allow_html=True,
                )

    remaining = [(label, value) for label, value in contract.items() if label not in shown_fields and value]
    if remaining:
        with st.expander("All other workbook fields"):
            for label, value in remaining:
                st.markdown(
                    f'<div class="detail-item"><span>{escape(label)}</span>'
                    f'<p>{escape(value).replace(chr(10), "<br>")}</p></div>',
                    unsafe_allow_html=True,
                )


st.set_page_config(
    page_title="Banedanmark railway map",
    page_icon="🛤️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown(
    """
    <style>
    /* Contract dashboard visual system v2 */
    :root {
        --panel: #0b1220;
        --panel-raised: #111c2e;
        --panel-soft: #172337;
        --border: rgba(148, 163, 184, 0.18);
        --text: #f8fafc;
        --muted: #9baabe;
        --accent: #f5b942;
        --accent-soft: rgba(245, 185, 66, 0.13);
    }
    html, body, [class*="st-"] {
        font-family: Inter, ui-sans-serif, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }
    header[data-testid="stHeader"] { display: none; }
    [data-testid="collapsedControl"] { display: none; }
    [data-testid="stAppViewContainer"] {
        background: var(--panel);
        color: var(--text);
    }
    [data-testid="stMainBlockContainer"] {
        max-width: none;
        padding: 0;
    }
    .stApp, .stApp > div { overflow: hidden; background: var(--panel); }
    [data-testid="stHorizontalBlock"] { gap: 0 !important; }
    [data-testid="stColumn"]:first-child {
        background:
            radial-gradient(circle at 0 0, rgba(245,185,66,.09), transparent 32%),
            var(--panel);
        border-right: 1px solid var(--border);
        padding: 34px 30px 28px 34px;
    }
    [data-testid="stColumn"]:last-child {
        background: #d8e7ea;
        padding: 12px 12px 12px 0;
    }
    [data-testid="stVerticalBlockBorderWrapper"] {
        scrollbar-color: #34445b transparent;
        scrollbar-width: thin;
    }
    iframe {
        display: block;
        border-radius: 14px;
        box-shadow: 0 18px 45px rgba(0, 0, 0, .25);
    }
    h1, h2, h3, p, label, [data-testid="stCaptionContainer"] {
        color: var(--text) !important;
    }
    .panel-kicker {
        color: var(--accent);
        font-size: 11px;
        font-weight: 800;
        letter-spacing: .15em;
        margin: 0 0 22px;
    }
    .route-header {
        border-bottom: 1px solid var(--border);
        margin-bottom: 18px;
        padding-bottom: 22px;
    }
    .route-header > span {
        color: var(--muted);
        font-size: 10px;
        font-weight: 800;
        letter-spacing: .12em;
    }
    .route-header h1 {
        color: var(--text);
        font-size: clamp(25px, 2vw, 36px);
        font-weight: 720;
        letter-spacing: -.035em;
        line-height: 1.08;
        margin: 8px 0 16px;
    }
    .route-tags { display: flex; flex-wrap: wrap; gap: 7px; }
    .route-tags b {
        background: var(--panel-soft);
        border: 1px solid var(--border);
        border-radius: 999px;
        color: #cbd5e1;
        font-size: 11px;
        font-weight: 650;
        padding: 6px 10px;
    }
    .match-count {
        align-items: center;
        color: var(--muted);
        display: flex;
        font-size: 12px;
        font-weight: 650;
        gap: 8px;
        margin: 4px 0 10px;
    }
    .match-count span {
        align-items: center;
        background: var(--accent);
        border-radius: 50%;
        color: #172033;
        display: inline-flex;
        font-weight: 850;
        height: 23px;
        justify-content: center;
        width: 23px;
    }
    [data-testid="stSelectbox"] label p {
        color: var(--muted) !important;
        font-size: 11px;
        font-weight: 750;
        letter-spacing: .07em;
        text-transform: uppercase;
    }
    [data-testid="stSelectbox"] [data-baseweb="select"] > div {
        background: var(--panel-raised);
        border-color: var(--border);
        border-radius: 10px;
        color: var(--text);
        min-height: 48px;
    }
    .fact-grid {
        display: grid;
        gap: 9px;
        grid-template-columns: repeat(2, minmax(0, 1fr));
        margin: 18px 0;
    }
    .fact-card {
        background: linear-gradient(145deg, rgba(255,255,255,.045), rgba(255,255,255,.018));
        border: 1px solid var(--border);
        border-radius: 11px;
        min-height: 76px;
        padding: 13px 14px;
    }
    .fact-card span, .detail-item span {
        color: var(--muted);
        display: block;
        font-size: 10px;
        font-weight: 800;
        letter-spacing: .06em;
        line-height: 1.35;
        margin-bottom: 7px;
        text-transform: uppercase;
    }
    .fact-card strong {
        color: var(--text);
        display: block;
        font-size: 14px;
        font-weight: 650;
        line-height: 1.35;
        overflow-wrap: anywhere;
    }
    [data-testid="stExpander"] {
        background: var(--panel-raised);
        border: 1px solid var(--border);
        border-radius: 11px;
        margin-bottom: 9px;
        overflow: hidden;
    }
    [data-testid="stExpander"] details summary {
        background: transparent !important;
        color: var(--text) !important;
        font-size: 13px;
        font-weight: 680;
        min-height: 51px;
    }
    [data-testid="stExpander"] details summary:hover {
        background: rgba(255,255,255,.035) !important;
    }
    [data-testid="stExpander"] svg { fill: var(--muted); }
    .detail-item {
        border-top: 1px solid var(--border);
        padding: 15px 2px 13px;
    }
    .detail-item:first-child { border-top: 0; padding-top: 4px; }
    .detail-item p {
        color: #dbe4ef !important;
        font-size: 13px;
        line-height: 1.62;
        margin: 0;
        overflow-wrap: anywhere;
    }
    .empty-state {
        align-items: flex-start;
        background: linear-gradient(145deg, rgba(255,255,255,.045), rgba(255,255,255,.015));
        border: 1px solid var(--border);
        border-radius: 16px;
        display: flex;
        flex-direction: column;
        margin-top: 28px;
        padding: 24px;
    }
    .empty-icon {
        align-items: center;
        background: var(--accent-soft);
        border: 1px solid rgba(245,185,66,.28);
        border-radius: 12px;
        color: var(--accent);
        display: flex;
        font-size: 22px;
        height: 45px;
        justify-content: center;
        margin-bottom: 19px;
        width: 45px;
    }
    .empty-state strong { color: var(--text); font-size: 17px; margin-bottom: 7px; }
    .empty-state span { color: var(--muted); font-size: 13px; line-height: 1.55; }
    [data-testid="stAlert"] { border-radius: 11px; }
    @media (max-width: 900px) {
        .stApp, .stApp > div { overflow: auto; }
        [data-testid="stColumn"]:first-child,
        [data-testid="stColumn"]:last-child { padding: 20px; }
        .fact-grid { grid-template-columns: 1fr; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

railway_geometry = load_railway_geometry()
contracts = load_contracts()
add_contract_tooltips(railway_geometry, contracts)


def selected_properties_from_state(state: object) -> dict | None:
    if not state:
        return None
    selection = state.get("selection", {})
    objects = selection.get("objects", {}).get("railway-stretches", [])
    if not objects:
        return None
    selected = objects[0]
    return selected.get("properties", selected)


previous_selection = selected_properties_from_state(
    st.session_state.get("railway-map", {})
)

railway_layer = pdk.Layer(
    "GeoJsonLayer",
    railway_geometry,
    id="railway-lines",
    stroked=True,
    filled=True,
    get_fill_color=[0, 0, 0, 42],
    get_line_color=[0, 0, 0, 255],
    get_line_width=2,
    line_width_min_pixels=2.5,
    line_width_max_pixels=5,
    pickable=False,
)

# The source shapes are very narrow polygons. This almost-transparent layer
# provides a generous hit area without making the rendered railway too thick.
railway_hover_layer = pdk.Layer(
    "GeoJsonLayer",
    railway_geometry,
    id="railway-stretches",
    stroked=True,
    filled=True,
    get_fill_color=[0, 0, 0, 1],
    get_line_color=[0, 0, 0, 1],
    get_line_width=8,
    line_width_units="pixels",
    line_width_min_pixels=8,
    line_width_max_pixels=10,
    pickable=True,
    auto_highlight=True,
    highlight_color=[255, 174, 0, 255],
)

map_layers = [railway_layer, railway_hover_layer]
if previous_selection:
    selected_id = previous_selection.get("STRKAFS")
    selected_features = [
        feature
        for feature in railway_geometry["features"]
        if str(feature["properties"].get("STRKAFS")) == str(selected_id)
    ]
    if selected_features:
        selected_layer = pdk.Layer(
            "GeoJsonLayer",
            {"type": "FeatureCollection", "features": selected_features},
            id="selected-railway-stretch",
            stroked=True,
            filled=True,
            get_fill_color=[255, 174, 0, 115],
            get_line_color=[255, 153, 0, 255],
            get_line_width=5,
            line_width_min_pixels=3,
            line_width_max_pixels=8,
            pickable=False,
        )
        map_layers.append(selected_layer)

deck = pdk.Deck(
    layers=map_layers,
    initial_view_state=pdk.ViewState(
        latitude=56.15,
        longitude=10.05,
        zoom=5.65,
        min_zoom=5,
        max_zoom=18,
    ),
    map_provider="maplibre",
    map_style="https://basemaps.cartocdn.com/gl/positron-gl-style/style.json",
    map_projection="globe",
    tooltip={
        "html": "<div style='padding:10px 13px;font-size:14px;font-weight:700;'>{HOVER_NAME}</div>",
        "style": {
            "backgroundColor": "rgba(15, 20, 27, 0.97)",
            "color": "#f8fafc",
            "padding": "0",
            "border": "1px solid rgba(132, 145, 163, 0.3)",
            "borderRadius": "9px",
            "boxShadow": "0 12px 30px rgba(0, 0, 0, 0.32)",
        },
    },
)
# Streamlit's Deck JSON converter requires a Deck expression here, not a
# JavaScript arrow function. This keeps the cursor pointy without parser errors.
deck.get_cursor = "@@='pointer'"

details_column, map_column = st.columns([0.38, 0.62], gap=None)
with details_column:
    details_panel = st.container(height=900, border=False)

with map_column:
    event = st.pydeck_chart(
        deck,
        height=900,
        width="stretch",
        on_select="rerun",
        selection_mode="single-object",
        key="railway-map",
    )

selected_objects = event.selection.objects.get("railway-stretches", [])
selected_object = selected_objects[0] if selected_objects else None
selected_properties = (
    selected_object.get("properties", selected_object) if selected_object else None
)
with details_panel:
    render_contract_panel(selected_properties, contracts)
