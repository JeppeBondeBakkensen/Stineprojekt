from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).parent
GEOJSON_FILE = ROOT / "BI_BTR_RELEASEAFSNIT_OD_1972479174689089199.geojson"
HOVER_FILE = ROOT / "railway_hover_names.xlsx"
STRETCH_FILE = ROOT / "railway_stretch_names.xlsx"
CONTRACT_FILE = next(ROOT.glob("Kopi af Oversigt*.xlsx"))
CONTRACT_SHEET = "Kopi af Oversigt over kontrakte"
OUTPUT_FILE = ROOT / "combined_railway_contracts.xlsx"


def text(value: object) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def normalized(value: object) -> str:
    return " ".join(
        unicodedata.normalize("NFKD", text(value))
        .casefold()
        .replace("–", "-")
        .replace("—", "-")
        .split()
    )


def load_names(path: Path) -> dict[int, str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Rail stretches"]
    names = {
        int(section_id): text(name)
        for section_id, name in worksheet.iter_rows(min_row=2, max_col=2, values_only=True)
        if section_id is not None and name
    }
    workbook.close()
    return names


def load_contracts() -> tuple[list[str], list[dict[str, str]]]:
    workbook = load_workbook(CONTRACT_FILE, read_only=True, data_only=True)
    worksheet = workbook[CONTRACT_SHEET]
    rows = worksheet.iter_rows(min_row=1, max_col=22, values_only=True)
    headers = [text(value) for value in next(rows)]
    contracts = []
    for row in rows:
        contract = {header: text(value) for header, value in zip(headers, row) if header}
        if contract.get("SAP nummer") or contract.get("Kontraktens titel"):
            contracts.append(contract)
    workbook.close()
    return headers, contracts


def route_codes(properties: dict) -> set[str]:
    codes: set[str] = set()
    for raw_code in re.findall(r"\d+", text(properties.get("TIB"))):
        number = int(raw_code)
        codes.add(str(number))
        if 1000 <= number <= 1099:
            codes.add(str(number - 1000))
        if 800 <= number <= 899 and number % 10 == 0:
            codes.add(str(number // 10))
    if properties.get("BANENR"):
        codes.add(text(properties["BANENR"]))
    return codes


def contract_matches(properties: dict, contract: dict[str, str]) -> bool:
    geography = normalized(
        contract.get(
            "Geografisk område + TIB-strækninger med responstid for enkelt strækning",
            "",
        )
    )
    if any(phrase in geography for phrase in ("hele landet", "landsdaekkende", "landsdækkende")):
        return True
    section_name = normalized(properties.get("NAVN"))
    if section_name and section_name in geography:
        return True
    return any(
        re.search(rf"(?<!\d){re.escape(code)}(?!\d)", geography)
        for code in route_codes(properties)
    )


def main() -> None:
    hover_names = load_names(HOVER_FILE)
    stretch_names = load_names(STRETCH_FILE)
    contract_headers, contracts = load_contracts()
    with GEOJSON_FILE.open(encoding="utf-8") as source:
        features = json.load(source)["features"]

    railway_headers = [
        "STRKAFS",
        "Railway name",
        "Hover name",
        "Stretch name",
        "FORKORTELSE",
        "FRA_KM",
        "TIL_KM",
        "Length km",
        "TIB",
        "BANENR",
        "BANENR_NAVN",
        "HVDSTRK",
        "HVDSTRK_NAVN",
        "DELSTRK",
        "DELSTRK_NAVN",
        "BTRTYPETEKST",
        "BANEKLASSETEKST",
        "Contract match",
    ]
    output_headers = railway_headers + [f"Contract: {header}" for header in contract_headers]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Railway and contracts"
    worksheet.append(output_headers)

    match_rows = 0
    unmatched_sections = 0
    for feature in features:
        properties = feature["properties"]
        section_id = int(properties["STRKAFS"])
        start_km = properties.get("FRA_KM")
        end_km = properties.get("TIL_KM")
        length_km = (
            round(abs(float(end_km) - float(start_km)), 3)
            if start_km is not None and end_km is not None
            else None
        )
        railway_values = [
            section_id,
            text(properties.get("NAVN")),
            hover_names.get(section_id, ""),
            stretch_names.get(section_id, ""),
            text(properties.get("FORKORTELSE")),
            start_km,
            end_km,
            length_km,
            text(properties.get("TIB")),
            text(properties.get("BANENR")),
            text(properties.get("BANENR_NAVN")),
            text(properties.get("HVDSTRK")),
            text(properties.get("HVDSTRK_NAVN")),
            text(properties.get("DELSTRK")),
            text(properties.get("DELSTRK_NAVN")),
            text(properties.get("BTRTYPETEKST")),
            text(properties.get("BANEKLASSETEKST")),
        ]
        matches = [contract for contract in contracts if contract_matches(properties, contract)]
        if not matches:
            unmatched_sections += 1
            worksheet.append(railway_values + ["No"] + [""] * len(contract_headers))
            continue
        for contract in matches:
            worksheet.append(
                railway_values
                + ["Yes"]
                + [contract.get(header, "") for header in contract_headers]
            )
            match_rows += 1

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 42
    for index, header in enumerate(output_headers, 1):
        width = 14
        if any(word in header for word in ("name", "NAVN", "titel", "område", "Beskrivelse")):
            width = 34
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(OUTPUT_FILE)
    print(f"Created: {OUTPUT_FILE.name}")
    print(f"Railway sections: {len(features)}")
    print(f"Contracts: {len(contracts)}")
    print(f"Matched section-contract rows: {match_rows}")
    print(f"Sections without a matched contract: {unmatched_sections}")
    print(f"Output data rows: {worksheet.max_row - 1}")


if __name__ == "__main__":
    main()
