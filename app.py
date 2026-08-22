from __future__ import annotations

import json
import math
from difflib import SequenceMatcher
from html import escape
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook
import pydeck as pdk
import streamlit as st


ROOT = Path(__file__).parent
GEOJSON_FILE = ROOT / "BI_BTR_RELEASEAFSNIT_OD_1972479174689089199.geojson"
OPTIMIZED_GEOJSON_FILE = ROOT / "railway_map_optimized.geojson"
GEOMETRY_CACHE_VERSION = 3
NAMES_FILE = ROOT / "railway_hover_names.xlsx"
CONTRACTS_FILE = next(ROOT.glob("Kopi af Oversigt*.xlsx"))
CONTRACT_SHEET = "Kopi af Oversigt over kontrakte"


def point_segment_distance(point: list[float], start: list[float], end: list[float]) -> float:
    """Return the planar distance used by the Ramer-Douglas-Peucker algorithm."""
    x, y = point
    start_x, start_y = start
    end_x, end_y = end
    delta_x = end_x - start_x
    delta_y = end_y - start_y
    if delta_x == 0 and delta_y == 0:
        return math.hypot(x - start_x, y - start_y)
    position = max(
        0.0,
        min(
            1.0,
            ((x - start_x) * delta_x + (y - start_y) * delta_y)
            / (delta_x * delta_x + delta_y * delta_y),
        ),
    )
    return math.hypot(
        x - (start_x + position * delta_x),
        y - (start_y + position * delta_y),
    )


def simplify_line(points: list[list[float]], tolerance: float) -> list[list[float]]:
    if len(points) <= 2:
        return points
    furthest_distance = 0.0
    furthest_index = 0
    for index, point in enumerate(points[1:-1], 1):
        distance = point_segment_distance(point, points[0], points[-1])
        if distance > furthest_distance:
            furthest_distance = distance
            furthest_index = index
    if furthest_distance <= tolerance:
        return [points[0], points[-1]]
    return (
        simplify_line(points[: furthest_index + 1], tolerance)[:-1]
        + simplify_line(points[furthest_index:], tolerance)
    )


def simplify_ring(ring: list[list[float]], tolerance: float = 0.0001) -> list[list[float]]:
    """Remove redundant low-detail vertices while retaining a valid closed polygon."""
    simplified = simplify_line(ring[:-1], tolerance)
    if len(simplified) < 3:
        return ring
    # Five decimal places is approximately metre-level precision in Denmark.
    # The source's 15 decimal places only inflate Streamlit's rerun payload.
    rounded = [[round(point[0], 5), round(point[1], 5)] for point in simplified]
    deduplicated = [rounded[0]]
    for point in rounded[1:]:
        if point != deduplicated[-1]:
            deduplicated.append(point)
    if len(deduplicated) < 3:
        return ring
    return deduplicated + [deduplicated[0]]


def simplify_geometry(geometry: dict) -> None:
    for feature in geometry["features"]:
        shape = feature["geometry"]
        if shape["type"] == "Polygon":
            shape["coordinates"] = [
                simplify_ring(ring) for ring in shape["coordinates"]
            ]
        elif shape["type"] == "MultiPolygon":
            shape["coordinates"] = [
                [simplify_ring(ring) for ring in polygon]
                for polygon in shape["coordinates"]
            ]


@st.cache_data
def load_railway_geometry(cache_version: int) -> dict:
    del cache_version  # Explicitly invalidates cached data when the asset schema changes.
    if OPTIMIZED_GEOJSON_FILE.exists():
        with OPTIMIZED_GEOJSON_FILE.open(encoding="utf-8") as source:
            return json.load(source)

    with GEOJSON_FILE.open(encoding="utf-8") as source:
        geometry = json.load(source)

    # The source contains many near-collinear points. Removing only redundant
    # sub-metre vertices substantially cuts the browser payload on every rerun.
    simplify_geometry(geometry)

    workbook = load_workbook(NAMES_FILE, read_only=True, data_only=True)
    worksheet = workbook["Rail stretches"]
    names = {
        int(stretch_id): str(name)
        for stretch_id, name in worksheet.iter_rows(min_row=2, values_only=True)
        if stretch_id is not None and name
    }
    workbook.close()

    for feature in geometry["features"]:
        properties = feature["properties"]
        # Only these values are needed client-side. The source has dozens of
        # additional attributes which otherwise get resent on every click.
        feature["properties"] = {
            "STRKAFS": properties.get("STRKAFS"),
            "NAVN": properties.get("NAVN"),
            "HOVER_NAME": names.get(
                int(properties["STRKAFS"]), properties.get("NAVN", "")
            ),
            "TIB": properties.get("TIB"),
            "BANENR": properties.get("BANENR"),
            "FORKORTELSE": properties.get("FORKORTELSE"),
            "BANENR_NAVN": properties.get("BANENR_NAVN"),
            "HVDSTRK_NAVN": properties.get("HVDSTRK_NAVN"),
            "DELSTRK_NAVN": properties.get("DELSTRK_NAVN"),
            "BTRTYPETEKST": properties.get("BTRTYPETEKST"),
            "BANEKLASSETEKST": properties.get("BANEKLASSETEKST"),
            "FRA_KM": properties.get("FRA_KM"),
            "TIL_KM": properties.get("TIL_KM"),
        }

    return geometry


def display_value(value: object) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


@st.cache_data
def load_contracts() -> list[dict[str, str]]:
    workbook = load_workbook(CONTRACTS_FILE, read_only=True, data_only=True)
    worksheet = workbook[CONTRACT_SHEET]
    rows = worksheet.iter_rows(min_row=1, max_col=22, values_only=True)
    headers = [display_value(value) for value in next(rows)]
    contracts = []
    for row in rows:
        contract = {
            header: display_value(value)
            for header, value in zip(headers, row)
            if header and display_value(value)
        }
        if contract.get("SAP nummer") or contract.get("Kontraktens titel"):
            contracts.append(contract)
    workbook.close()
    return contracts


def normalized_text(value: object) -> str:
    text = "".join(
        character
        for character in unicodedata.normalize("NFKD", display_value(value)).casefold()
        if not unicodedata.combining(character)
    )
    text = re.sub(r"\barhus\b", "aarhus", text)
    text = re.sub(r"\balborg\b", "aalborg", text)
    return " ".join(text.replace("–", "-").replace("—", "-").split())


def route_codes(properties: dict) -> set[str]:
    codes = set()
    for raw_code in re.findall(r"\d+", display_value(properties.get("TIB"))):
        number = int(raw_code)
        codes.add(str(number))
        if 1000 <= number <= 1099:
            codes.add(str(number - 1000))
        if 800 <= number <= 899 and number % 10 == 0:
            codes.add(str(number // 10))
    if properties.get("BANENR"):
        codes.add(display_value(properties["BANENR"]))
    return codes


def matching_contracts(properties: dict, contracts: list[dict[str, str]]) -> list[dict[str, str]]:
    stretch_name = normalized_text(properties.get("NAVN"))
    codes = route_codes(properties)
    matches = []
    for contract in contracts:
        geography = normalized_text(
            contract.get(
                "Geografisk område + TIB-strækninger med responstid for enkelt strækning",
                "",
            )
        )
        nationwide = any(
            phrase in geography
            for phrase in ("hele landet", "landsdaekkende", "landsdækkende")
        )
        named_match = stretch_name and stretch_name in geography
        code_match = any(
            re.search(rf"(?<!\d){re.escape(code)}(?!\d)", geography)
            for code in codes
        )
        if nationwide or named_match or code_match:
            matches.append(contract)
    return matches


def tooltip_text(value: object, limit: int = 180) -> str:
    text = " ".join(display_value(value).split())
    if len(text) > limit:
        text = f"{text[: limit - 1].rstrip()}…"
    return escape(text or "—")


def add_contract_tooltips(geometry: dict, contracts: list[dict[str, str]]) -> None:
    for feature in geometry["features"]:
        properties = feature["properties"]
        matches = matching_contracts(properties, contracts)
        start_km = properties.get("FRA_KM")
        end_km = properties.get("TIL_KM")
        length_km = (
            f"{abs(float(end_km) - float(start_km)):.1f} km"
            if start_km is not None and end_km is not None
            else "—"
        )
        properties["TOOLTIP_STRETCH"] = tooltip_text(properties.get("HOVER_NAME"))
        properties["TOOLTIP_CONTEXT"] = tooltip_text(
            properties.get("DELSTRK_NAVN") or properties.get("HVDSTRK_NAVN")
        )
        properties["TOOLTIP_LENGTH"] = length_km
        properties["TOOLTIP_TYPE"] = tooltip_text(properties.get("BTRTYPETEKST"), 50)
        properties["TOOLTIP_CLASS"] = tooltip_text(
            properties.get("BANEKLASSETEKST"), 50
        )
        properties["CONTRACT_COUNT"] = len(matches)
        if not matches:
            properties.update(
                {
                    "CONTRACT_TITLE": "No matching contract",
                    "CONTRACT_SUPPLIER": "—",
                    "CONTRACT_SAP": "—",
                    "CONTRACT_RESPONSE": "—",
                    "CONTRACT_MORE": "Click the stretch to inspect its route data",
                }
            )
            continue

        primary = matches[0]
        additional = len(matches) - 1
        properties.update(
            {
                "CONTRACT_TITLE": tooltip_text(primary.get("Kontraktens titel")),
                "CONTRACT_SUPPLIER": tooltip_text(primary.get("Leverandørnavn")),
                "CONTRACT_SAP": tooltip_text(primary.get("SAP nummer"), 60),
                "CONTRACT_RESPONSE": tooltip_text(primary.get("Responstid")),
                "CONTRACT_MORE": (
                    f"+{additional} additional contract{'s' if additional != 1 else ''} · click for details"
                    if additional
                    else "Click for full contract details"
                ),
            }
        )


def render_contract_panel(properties: dict | None, contracts: list[dict[str, str]]) -> None:
    st.markdown(
        '<div class="panel-kicker">BANEDANMARK · CONTRACT MAP</div>',
        unsafe_allow_html=True,
    )
    if not properties:
        st.markdown("# Select a railway stretch")
        st.markdown(
            '<div class="empty-state">'
            '<div class="empty-icon">↖</div>'
            '<strong>Explore the railway map</strong>'
            '<span>Click a black railway stretch to view matching contracts, '
            'response times and operational requirements.</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        return

    route_name = escape(display_value(properties.get("NAVN")) or "Selected stretch")
    st.markdown(
        f'<div class="route-header"><span>SELECTED STRETCH</span>'
        f'<h1>{route_name}</h1>'
        f'<div class="route-tags"><b>TIB/BTR {escape(display_value(properties.get("TIB")) or "—")}</b>'
        f'<b>Railway {escape(display_value(properties.get("BANENR")) or "—")}</b>'
        f'<b>km {escape(display_value(properties.get("FRA_KM")) or "—")}–'
        f'{escape(display_value(properties.get("TIL_KM")) or "—")}</b></div></div>',
        unsafe_allow_html=True,
    )
    matches = matching_contracts(properties, contracts)
    if not matches:
        st.warning("No matching contract was found in the contract workbook.")
        return

    st.markdown(
        f'<div class="match-count"><span>{len(matches)}</span> matching contract'
        f'{"s" if len(matches) != 1 else ""}</div>',
        unsafe_allow_html=True,
    )
    selected_index = st.selectbox(
        "Active contract",
        range(len(matches)),
        format_func=lambda index: (
            f"{matches[index].get('Kontraktens titel', 'Contract')} · "
            f"{matches[index].get('SAP nummer', 'No SAP number')}"
        ),
        key=f"contract-{properties.get('STRKAFS', 'selected')}",
    )
    contract = matches[selected_index]

    def fact_card(label: str, value: str) -> str:
        return (
            '<div class="fact-card">'
            f'<span>{escape(label)}</span><strong>{escape(value or "—")}</strong>'
            '</div>'
        )

    st.markdown(
        '<div class="fact-grid">'
        + fact_card("Supplier", contract.get("Leverandørnavn", ""))
        + fact_card("SAP number", contract.get("SAP nummer", ""))
        + fact_card("Effective from", contract.get("Ikraft-trædelse", ""))
        + fact_card(
            "Duration",
            contract.get(
                "Varighed \ni år (inkl mobilisering + option på forlængelse)", ""
            ),
        )
        + '</div>',
        unsafe_allow_html=True,
    )

    sections = [
        (
            "Coverage & response",
            [
                "Geografisk område + TIB-strækninger med responstid for enkelt strækning",
                "Responstid",
                "Beskrivelse af beredskabet",
            ],
        ),
        (
            "Capacity & operations",
            [
                "Antal fejl, som leverandøren skal kunne håndtere på én gang.",
                "Antal arbejdshold og personer pr. hold, som leverandøren har pligt til at stille med.",
                "Beskrevet fremgangsmåde omkring fejlretning",
                "Krav til data registrering",
                "Afrapporteringsfrekvens for leverandør",
            ],
        ),
        (
            "Penalties & obligations",
            [
                "Oversigt over bodsbelagte emner",
                "Beskrivelse af og i givet fald hvordan bod/boderne maksimeres",
                "Beskrivelse af boderne",
                "Henvisning til hvor der er beskrevet i kontrakten",
                "Henvisning til hvor der er beskrevet i kontrakten2",
            ],
        ),
        (
            "Ownership & definitions",
            ["Vedligeholdelsesforvalter", "GFS", "Kontraktansvarlig (leder)", "Definitioner"],
        ),
    ]
    shown_fields = {
        "Kontraktens titel",
        "SAP nummer",
        "Leverandørnavn",
        "Ikraft-trædelse",
        "Varighed \ni år (inkl mobilisering + option på forlængelse)",
    }
    for title, fields in sections:
        populated = [(field, contract.get(field, "")) for field in fields if contract.get(field)]
        shown_fields.update(fields)
        if not populated:
            continue
        with st.expander(title, expanded=title == "Coverage & response"):
            for label, value in populated:
                st.markdown(
                    f'<div class="detail-item"><span>{escape(label)}</span>'
                    f'<p>{escape(value).replace(chr(10), "<br>")}</p></div>',
                    unsafe_allow_html=True,
                )

    remaining = [(label, value) for label, value in contract.items() if label not in shown_fields and value]
    if remaining:
        with st.expander("All other workbook fields"):
            for label, value in remaining:
                st.markdown(
                    f'<div class="detail-item"><span>{escape(label)}</span>'
                    f'<p>{escape(value).replace(chr(10), "<br>")}</p></div>',
                    unsafe_allow_html=True,
                )


st.set_page_config(
    page_title="Banedanmark railway map",
    page_icon="🛤️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown(
    """
    <style>
    /* Contract dashboard visual system v2 */
    :root {
        --panel: #0b1220;
        --panel-raised: #111c2e;
        --panel-soft: #172337;
        --border: rgba(148, 163, 184, 0.18);
        --text: #f8fafc;
        --muted: #9baabe;
        --accent: #f5b942;
        --accent-soft: rgba(245, 185, 66, 0.13);
    }
    html, body, [class*="st-"] {
        font-family: Inter, ui-sans-serif, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }
    span[data-testid="stIconMaterial"] {
        font-family: "Material Symbols Rounded", "Material Symbols Outlined" !important;
    }
    header[data-testid="stHeader"] { display: none; }
    [data-testid="collapsedControl"] { display: none; }
    [data-testid="stAppViewContainer"] {
        background: var(--panel);
        color: var(--text);
    }
    [data-testid="stMainBlockContainer"] {
        max-width: none;
        padding: 0;
    }
    .stApp, .stApp > div { overflow: hidden; background: var(--panel); }
    [data-testid="stHorizontalBlock"] { gap: 0 !important; }
    [data-testid="stColumn"]:first-child {
        background:
            radial-gradient(circle at 0 0, rgba(245,185,66,.09), transparent 32%),
            var(--panel);
        border-right: 1px solid var(--border);
        padding: 34px 30px 28px 34px;
    }
    [data-testid="stColumn"]:last-child {
        background: var(--panel);
        padding: 12px 12px 12px 0;
    }
    [data-testid="stVerticalBlockBorderWrapper"] {
        scrollbar-color: #34445b transparent;
        scrollbar-width: thin;
    }
    iframe {
        display: block;
        border-radius: 12px;
        box-shadow: 0 1px 3px rgba(16,24,40,.10);
    }
    h1, h2, h3, p, label, [data-testid="stCaptionContainer"] {
        color: var(--text) !important;
    }
    .panel-kicker {
        color: var(--accent);
        font-size: 11px;
        font-weight: 800;
        letter-spacing: .15em;
        margin: 0 0 22px;
    }
    .route-header {
        border-bottom: 1px solid var(--border);
        margin-bottom: 18px;
        padding-bottom: 22px;
    }
    .route-header > span {
        color: var(--muted);
        font-size: 10px;
        font-weight: 800;
        letter-spacing: .12em;
    }
    .route-header h1 {
        color: var(--text);
        font-size: clamp(25px, 2vw, 36px);
        font-weight: 720;
        letter-spacing: -.035em;
        line-height: 1.08;
        margin: 8px 0 16px;
    }
    .route-tags { display: flex; flex-wrap: wrap; gap: 7px; }
    .route-tags b {
        background: var(--panel-soft);
        border: 1px solid var(--border);
        border-radius: 999px;
        color: #cbd5e1;
        font-size: 11px;
        font-weight: 650;
        padding: 6px 10px;
    }
    .match-count {
        align-items: center;
        color: var(--muted);
        display: flex;
        font-size: 12px;
        font-weight: 650;
        gap: 8px;
        margin: 4px 0 10px;
    }
    .match-count span {
        align-items: center;
        background: var(--accent);
        border-radius: 50%;
        color: #172033;
        display: inline-flex;
        font-weight: 850;
        height: 23px;
        justify-content: center;
        width: 23px;
    }
    [data-testid="stSelectbox"] label p {
        color: var(--muted) !important;
        font-size: 11px;
        font-weight: 750;
        letter-spacing: .07em;
        text-transform: uppercase;
    }
    [data-testid="stSelectbox"] [data-baseweb="select"] > div {
        background: var(--panel-raised);
        border-color: var(--border);
        border-radius: 10px;
        color: var(--text);
        min-height: 48px;
    }
    .fact-grid {
        display: grid;
        gap: 9px;
        grid-template-columns: repeat(2, minmax(0, 1fr));
        margin: 18px 0;
    }
    .fact-card {
        background: linear-gradient(145deg, rgba(255,255,255,.045), rgba(255,255,255,.018));
        border: 1px solid var(--border);
        border-radius: 11px;
        min-height: 76px;
        padding: 13px 14px;
    }
    .fact-card span, .detail-item span {
        color: var(--muted);
        display: block;
        font-size: 10px;
        font-weight: 800;
        letter-spacing: .06em;
        line-height: 1.35;
        margin-bottom: 7px;
        text-transform: uppercase;
    }
    .fact-card strong {
        color: var(--text);
        display: block;
        font-size: 14px;
        font-weight: 650;
        line-height: 1.35;
        overflow-wrap: anywhere;
    }
    [data-testid="stExpander"] {
        background: var(--panel-raised);
        border: 1px solid var(--border);
        border-radius: 11px;
        margin-bottom: 9px;
        overflow: hidden;
    }
    [data-testid="stExpander"] details summary {
        background: transparent !important;
        color: var(--text) !important;
        font-size: 13px;
        font-weight: 680;
        min-height: 51px;
    }
    [data-testid="stExpander"] details summary:hover {
        background: rgba(255,255,255,.035) !important;
    }
    [data-testid="stExpander"] svg { fill: var(--muted); }
    .detail-item {
        border-top: 1px solid var(--border);
        padding: 15px 2px 13px;
    }
    .detail-item:first-child { border-top: 0; padding-top: 4px; }
    .detail-item p {
        color: #dbe4ef !important;
        font-size: 13px;
        line-height: 1.62;
        margin: 0;
        overflow-wrap: anywhere;
    }
    .empty-state {
        align-items: flex-start;
        background: linear-gradient(145deg, rgba(255,255,255,.045), rgba(255,255,255,.015));
        border: 1px solid var(--border);
        border-radius: 16px;
        display: flex;
        flex-direction: column;
        margin-top: 28px;
        padding: 24px;
    }
    .empty-icon {
        align-items: center;
        background: var(--accent-soft);
        border: 1px solid rgba(245,185,66,.28);
        border-radius: 12px;
        color: var(--accent);
        display: flex;
        font-size: 22px;
        height: 45px;
        justify-content: center;
        margin-bottom: 19px;
        width: 45px;
    }
    .empty-state strong { color: var(--text); font-size: 17px; margin-bottom: 7px; }
    .empty-state span { color: var(--muted); font-size: 13px; line-height: 1.55; }
    .filter-kicker {
        color: var(--accent);
        font-size: 10px;
        font-weight: 850;
        letter-spacing: .14em;
        margin: 0 0 10px;
    }
    [data-testid="stTextInput"] input {
        background: var(--panel-raised);
        border-color: var(--border);
        border-radius: 11px;
        color: var(--text);
        min-height: 46px;
    }
    [data-testid="stTextInput"] input:focus { border-color: rgba(245,185,66,.65); }
    [data-testid="stMultiSelect"] [data-baseweb="select"] > div {
        background: var(--panel);
        border-color: var(--border);
        border-radius: 9px;
        color: var(--text);
    }
    [data-testid="stMultiSelect"] label p {
        color: var(--muted) !important;
        font-size: 11px;
        font-weight: 700;
    }
    [data-baseweb="popover"], [data-baseweb="menu"] {
        background: #172337 !important;
        color: #f8fafc !important;
    }
    [data-baseweb="popover"] li, [role="option"] {
        background: #172337 !important;
        color: #f8fafc !important;
    }
    [data-baseweb="popover"] li:hover, [role="option"]:hover {
        background: #25344a !important;
    }
    [data-testid="stExpander"] details summary p {
        color: #f8fafc !important;
        font-weight: 750;
    }
    .filter-results {
        color: var(--muted);
        font-size: 11px;
        margin: 8px 0 16px;
    }
    .filter-results strong { color: var(--accent); font-size: 13px; }
    .st-key-map-toolbar {
        background: transparent;
        border-radius: 0;
        box-shadow: none;
        margin: 0;
        padding: 0 0 8px;
        position: relative;
        z-index: 20;
    }
    .st-key-map-toolbar [data-testid="stHorizontalBlock"] {
        align-items: start;
        display: grid !important;
        gap: 10px !important;
        grid-template-columns: minmax(0, 1fr) auto;
    }
    .st-key-map-toolbar [data-testid="stColumn"] {
        background: transparent !important;
        border: 0 !important;
        flex: none !important;
        padding: 0 !important;
        width: auto !important;
    }
    .st-key-map-toolbar [data-testid="stVerticalBlock"] { gap: 0 !important; }
    .st-key-map-toolbar iframe {
        background: #111c2e;
        border: 1px solid #334155;
        border-radius: 8px;
        box-shadow: none;
        height: 42px !important;
    }
    .st-key-map-toolbar [data-testid="stTextInput"] div[data-baseweb="input"] {
        background: #ffffff;
        border: 1px solid #cbd5e1;
        border-radius: 8px;
        box-shadow: none;
        min-height: 42px;
        overflow: hidden;
    }
    .st-key-map-toolbar [data-testid="stTextInput"] div[data-baseweb="input"]:focus-within {
        border-color: #94a3b8;
        box-shadow: 0 0 0 3px rgba(17,24,39,.12);
    }
    .st-key-map-toolbar [data-testid="stTextInput"] input {
        background-color: transparent !important;
        background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='16' height='16' viewBox='0 0 24 24' fill='none' stroke='%2364758b' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Ccircle cx='11' cy='11' r='8'/%3E%3Cpath d='m21 21-4.35-4.35'/%3E%3C/svg%3E");
        background-position: 13px center;
        background-repeat: no-repeat;
        background-size: 16px;
        border: 0 !important;
        color: #0f172a;
        min-height: 40px;
        padding-left: 40px;
    }
    .st-key-map-toolbar [data-testid="stTextInput"] input::placeholder {
        color: #64748b;
        opacity: 1;
    }
    [data-testid="stColumn"]:has(.st-key-map-toolbar) { position: relative; }
    .st-key-search-suggestions {
        background: #111c2e;
        border: 1px solid #334155;
        border-radius: 8px;
        box-shadow: 0 12px 28px rgba(2,6,23,.28);
        left: 12px;
        margin: 0;
        height: min(320px, calc(100vh - 190px));
        max-height: 320px;
        overflow-x: hidden !important;
        overflow-y: scroll !important;
        overscroll-behavior: contain;
        padding: 4px;
        position: absolute;
        right: 140px;
        top: 60px;
        z-index: 1000;
        scrollbar-color: #475569 transparent;
        scrollbar-gutter: stable;
        scrollbar-width: thin;
    }
    .st-key-search-suggestions > div,
    .st-key-search-suggestions [data-testid="stVerticalBlock"] {
        min-height: max-content;
        overflow: visible !important;
    }
    .st-key-search-suggestions [data-testid="stRadio"] {
        margin: 0;
    }
    .st-key-search-suggestions [data-testid="stRadio"] div[role="radiogroup"] {
        gap: 2px;
    }
    .st-key-search-suggestions [data-testid="stRadio"] label {
        border-radius: 6px;
        margin: 0;
        min-height: 36px;
        padding: 7px 9px;
        width: 100%;
    }
    .st-key-search-suggestions [data-testid="stRadio"] label:hover {
        background: #1e2c41;
    }
    .st-key-search-suggestions [data-testid="stRadio"] label:has(input:checked) {
        background: rgba(245,185,66,.18);
    }
    .st-key-search-suggestions [data-testid="stRadio"] label p {
        color: #f8fafc !important;
        font-size: 12px;
        line-height: 1.3;
    }
    .st-key-search-suggestions [data-testid="stRadio"] label:has(input:checked) p {
        color: #f8c95f !important;
    }
    .st-key-map-toolbar [data-testid="stPopover"] button {
        background: var(--panel-raised);
        border: 1px solid #475569;
        border-radius: 8px;
        box-shadow: none;
        color: #f8fafc;
        font-size: 13px;
        font-weight: 700;
        min-height: 42px;
        width: 100%;
    }
    .st-key-map-toolbar [data-testid="stPopover"] { width: 118px; }
    .st-key-map-toolbar [data-testid="stPopover"] button:hover {
        background: #1e2c41;
        border-color: #64748b;
    }
    .map-filter-status {
        align-items: center;
        display: flex;
        flex-wrap: wrap;
        gap: 8px 12px;
        min-height: 20px;
        padding: 8px 2px 0;
    }
    .map-filter-status > span {
        color: var(--muted);
        font-size: 11px;
        white-space: nowrap;
    }
    .map-filter-status strong { color: var(--text); font-size: 12px; }
    .filter-chips { display: flex; flex-wrap: wrap; gap: 5px; }
    .filter-chip {
        background: rgba(245,185,66,.2);
        border: 1px solid rgba(189,126,0,.32);
        border-radius: 999px;
        color: #6f4900;
        font-size: 10px;
        font-weight: 750;
        padding: 3px 8px;
    }
    .popover-title {
        color: var(--accent);
        font-size: 10px;
        font-weight: 850;
        letter-spacing: .12em;
        margin-bottom: 10px;
    }
    [data-testid="stButton"] button {
        background: var(--panel-soft);
        border-color: var(--border);
        color: var(--text);
    }
    [data-testid="stAlert"] { border-radius: 11px; }
    @media (max-width: 900px) {
        .stApp, .stApp > div { overflow: auto; }
        [data-testid="stColumn"]:first-child,
        [data-testid="stColumn"]:last-child { padding: 20px; }
        .fact-grid { grid-template-columns: 1fr; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

contracts = load_contracts()
railway_geometry = load_railway_geometry(GEOMETRY_CACHE_VERSION)


def selected_properties_from_state(state: object) -> dict | None:
    if not state:
        return None
    selection = state.get("selection", {})
    objects = selection.get("objects", {}).get("railway-stretches", [])
    if not objects:
        return None
    selected = objects[0]
    return selected.get("properties", selected)


def render_search_entity_panel(entity: dict, contract_rows: list[dict[str, str]]) -> None:
    st.markdown('<div class="panel-kicker">SEARCH RESULT</div>', unsafe_allow_html=True)
    st.markdown(f'# {escape(entity["label"])}')
    if entity["kind"] == "Main route":
        st.markdown(
            f'<div class="match-count"><span>{len(entity["section_ids"])}</span> '
            'sections highlighted</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            '<div class="empty-state"><strong>Select the affected section</strong>'
            '<span>Click the relevant orange section on the map to see its matching '
            'contract, responsible people and response time.</span></div>',
            unsafe_allow_html=True,
        )
        return
    if entity["kind"] == "Contract":
        contract = contract_rows[entity["contract_index"]]
        contact_fields = (
            ("Supplier", contract.get("Leverandørnavn", "")),
            ("Maintenance manager", contract.get("Vedligeholdelsesforvalter", "")),
            ("Contract manager", contract.get("Kontraktansvarlig (leder)", "")),
            ("GFS", contract.get("GFS", "")),
            ("SAP number", contract.get("SAP nummer", "")),
            ("Response time", contract.get("Responstid", "")),
        )
        st.markdown(
            '<div class="match-count"><span>✓</span> matching railway sections highlighted</div>',
            unsafe_allow_html=True,
        )
        for label, value in contact_fields:
            if value:
                st.markdown(
                    f'<div class="detail-item"><span>{escape(label)}</span>'
                    f'<p>{escape(value).replace(chr(10), "<br>")}</p></div>',
                    unsafe_allow_html=True,
                )


previous_selection = selected_properties_from_state(
    st.session_state.get("railway-map", {})
)


def available_values(field: str) -> list[str]:
    return sorted(
        {
            display_value(feature["properties"].get(field))
            for feature in railway_geometry["features"]
            if display_value(feature["properties"].get(field))
        }
    )


@st.cache_data
def build_search_entities(cache_version: int) -> list[dict]:
    geometry = load_railway_geometry(cache_version)
    contract_rows = load_contracts()
    entities: list[dict] = []
    corridors: dict[str, list[dict]] = {}

    for feature in geometry["features"]:
        properties = feature["properties"]
        section_id = int(properties["STRKAFS"])
        codes = " ".join(sorted(route_codes(properties)))
        section_search = " ".join(
            display_value(properties.get(field))
            for field in (
                "NAVN", "HOVER_NAME", "FORKORTELSE", "STRKAFS", "TIB", "BANENR",
                "BANENR_NAVN", "HVDSTRK_NAVN", "DELSTRK_NAVN",
                "BTRTYPETEKST", "BANEKLASSETEKST",
            )
        )
        entities.append(
            {
                "id": f"section:{section_id}",
                "kind": "Section",
                "label": display_value(properties.get("HOVER_NAME") or properties.get("NAVN")),
                "subtitle": f'{display_value(properties.get("FORKORTELSE"))} · TIB {display_value(properties.get("TIB"))}',
                "section_ids": [section_id],
                "search_text": normalized_text(f"{section_search} {codes}"),
            }
        )
        corridor = display_value(properties.get("HVDSTRK_NAVN"))
        if corridor:
            corridors.setdefault(corridor, []).append(properties)

    for corridor, sections in corridors.items():
        section_ids = [int(section["STRKAFS"]) for section in sections]
        context = " ".join(
            " ".join(
                (
                    display_value(section.get("DELSTRK_NAVN")),
                    display_value(section.get("BANENR_NAVN")),
                    display_value(section.get("TIB")),
                    " ".join(sorted(route_codes(section))),
                )
            )
            for section in sections
        )
        entities.append(
            {
                "id": f"route:{corridor}",
                "kind": "Main route",
                "label": corridor,
                "subtitle": f"{len(section_ids)} sections",
                "section_ids": section_ids,
                "search_text": normalized_text(f"{corridor} {context} main route hovedstrækning"),
            }
        )

    for index, contract in enumerate(contract_rows):
        section_ids = [
            int(feature["properties"]["STRKAFS"])
            for feature in geometry["features"]
            if contract in matching_contracts(feature["properties"], contract_rows)
        ]
        searchable = " ".join(contract.values())
        entities.append(
            {
                "id": f"contract:{index}",
                "kind": "Contract",
                "label": contract.get("Kontraktens titel", "Contract"),
                "subtitle": " · ".join(
                    value
                    for value in (contract.get("Leverandørnavn"), contract.get("SAP nummer"))
                    if value
                ),
                "section_ids": section_ids,
                "contract_index": index,
                "search_text": normalized_text(searchable),
            }
        )
    return entities


def ranked_search(query: str, entities: list[dict], limit: int = 100) -> list[dict]:
    normalized_query = normalized_text(query)
    if not normalized_query:
        return []
    ignored = {
        "a", "an", "and", "at", "from", "in", "issue", "on", "the", "to",
        "en", "et", "fra", "i", "mellem", "og", "pa", "problem", "til",
    }
    query_tokens = [token for token in re.findall(r"[\wæøå-]+", normalized_query) if token not in ignored]
    scored = []
    for entity in entities:
        haystack = entity["search_text"]
        label = normalized_text(entity["label"])
        words = set(re.findall(r"[\wæøå-]+", haystack))
        exact_tokens = sum(token in haystack for token in query_tokens)
        fuzzy_tokens = sum(
            max((SequenceMatcher(None, token, word).ratio() for word in words), default=0) >= 0.78
            for token in query_tokens
            if token not in haystack
        )
        coverage = (exact_tokens + 0.65 * fuzzy_tokens) / max(len(query_tokens), 1)
        score = coverage * 70 + SequenceMatcher(None, normalized_query, label).ratio() * 25
        if normalized_query in haystack:
            score += 35
        if entity["kind"] == "Main route" and any(
            phrase in normalized_query for phrase in ("hovedstrækning", "main route", "route")
        ):
            score += 18
        if entity["kind"] == "Contract" and any(
            phrase in normalized_query for phrase in ("contact", "contract", "supplier", "kontakt", "leverandør")
        ):
            score += 15
        if score >= 38:
            scored.append((score, entity))
    scored.sort(key=lambda item: (-item[0], item[1]["kind"], item[1]["label"]))
    return [entity for _, entity in scored[:limit]]


def clear_map_filters() -> None:
    for key in (
        "rail-types", "rail-classes", "rail-tibs", "rail-numbers",
    ):
        st.session_state.pop(key, None)


selected_search_entity = None
details_column, map_column = st.columns([0.38, 0.62], gap=None)
with details_column:
    details_panel = st.container(height=900, border=False)

with map_column:
    with st.container(key="map-toolbar"):
        selected_filter_count = sum(
            bool(st.session_state.get(key))
            for key in ("rail-types", "rail-classes", "rail-tibs", "rail-numbers")
        )
        filter_label = (
            f"Filters · {selected_filter_count}" if selected_filter_count else "Filters"
        )
        with st.popover(filter_label, use_container_width=False):
            st.markdown('<div class="popover-title">FILTER THE NETWORK</div>', unsafe_allow_html=True)
            selected_types = st.multiselect(
                "Section type", available_values("BTRTYPETEKST"), key="rail-types"
            )
            selected_classes = st.multiselect(
                "Railway class", available_values("BANEKLASSETEKST"), key="rail-classes"
            )
            selected_tibs = st.multiselect(
                "TIB", available_values("TIB"), key="rail-tibs"
            )
            selected_railways = st.multiselect(
                "Railway number", available_values("BANENR"), key="rail-numbers"
            )
            st.button(
                "Clear all filters",
                on_click=clear_map_filters,
                use_container_width=True,
            )
        filter_status = st.empty()
    selected_section_ids = None
    filtered_features = []
    for feature in railway_geometry["features"]:
        properties = feature["properties"]
        if selected_section_ids is not None and int(properties["STRKAFS"]) not in selected_section_ids:
            continue
        if selected_types and display_value(properties.get("BTRTYPETEKST")) not in selected_types:
            continue
        if selected_classes and display_value(properties.get("BANEKLASSETEKST")) not in selected_classes:
            continue
        if selected_tibs and display_value(properties.get("TIB")) not in selected_tibs:
            continue
        if selected_railways and display_value(properties.get("BANENR")) not in selected_railways:
            continue
        filtered_features.append(feature)

    filters_active = bool(
        selected_search_entity
        or selected_types
        or selected_classes
        or selected_tibs
        or selected_railways
    )

    active_filter_labels = []
    if selected_search_entity:
        active_filter_labels.append(
            f'{escape(selected_search_entity["kind"])}: '
            f'{escape(selected_search_entity["label"])}'
        )
    for label, values in (
        ("Type", selected_types),
        ("Class", selected_classes),
        ("TIB", selected_tibs),
        ("Railway", selected_railways),
    ):
        active_filter_labels.extend(f"{label}: {escape(value)}" for value in values)
    chips = "".join(f'<span class="filter-chip">{label}</span>' for label in active_filter_labels)
    filter_status.markdown(
        '<div class="map-filter-status">'
        f'<span><strong>{len(filtered_features)}</strong> of '
        f'{len(railway_geometry["features"])} sections</span>'
        f'<div class="filter-chips">{chips}</div></div>',
        unsafe_allow_html=True,
    )

filtered_geometry = {"type": "FeatureCollection", "features": filtered_features}

railway_layer = pdk.Layer(
    "GeoJsonLayer",
    railway_geometry,
    id="railway-stretches",
    stroked=True,
    filled=True,
    get_fill_color=([45, 54, 64, 95] if filters_active else [0, 0, 0, 190]),
    get_line_color=([45, 54, 64, 150] if filters_active else [0, 0, 0, 255]),
    get_line_width=2.5,
    line_width_units="pixels",
    line_width_min_pixels=2.5,
    line_width_max_pixels=4,
    pickable=True,
    auto_highlight=True,
    highlight_color=[255, 174, 0, 255],
)

map_layers = [railway_layer]
if filters_active and filtered_features:
    filter_highlight_layer = pdk.Layer(
        "GeoJsonLayer",
        filtered_geometry,
        id="filtered-railway-stretches",
        stroked=True,
        filled=True,
        get_fill_color=[255, 174, 0, 190],
        get_line_color=[255, 153, 0, 255],
        get_line_width=3,
        line_width_units="pixels",
        line_width_min_pixels=3,
        line_width_max_pixels=5,
        pickable=False,
    )
    map_layers.append(filter_highlight_layer)

if previous_selection:
    selected_id = previous_selection.get("STRKAFS")
    selected_features = [
        feature
        for feature in railway_geometry["features"]
        if str(feature["properties"].get("STRKAFS")) == str(selected_id)
    ]
    if selected_features:
        selected_layer = pdk.Layer(
            "GeoJsonLayer",
            {"type": "FeatureCollection", "features": selected_features},
            id="selected-railway-stretch",
            stroked=True,
            filled=True,
            get_fill_color=[255, 174, 0, 115],
            get_line_color=[255, 153, 0, 255],
            get_line_width=3,
            line_width_min_pixels=2,
            line_width_max_pixels=4,
            pickable=False,
        )
        map_layers.append(selected_layer)

deck = pdk.Deck(
    layers=map_layers,
    initial_view_state=pdk.ViewState(
        latitude=56.15,
        longitude=10.05,
        zoom=5.65,
        min_zoom=5,
        max_zoom=18,
    ),
    map_provider="maplibre",
    map_style="https://basemaps.cartocdn.com/gl/positron-gl-style/style.json",
    map_projection="globe",
    tooltip={
        "html": "<div style='padding:10px 13px;font-size:14px;font-weight:700;'>{HOVER_NAME}</div>",
        "style": {
            "backgroundColor": "rgba(15, 20, 27, 0.97)",
            "color": "#f8fafc",
            "padding": "0",
            "border": "1px solid rgba(132, 145, 163, 0.3)",
            "borderRadius": "9px",
            "boxShadow": "0 12px 30px rgba(0, 0, 0, 0.32)",
        },
    },
)
# Streamlit's Deck JSON converter requires a Deck expression here, not a
# JavaScript arrow function. This keeps the cursor pointy without parser errors.
deck.get_cursor = "@@='pointer'"

with map_column:
    event = st.pydeck_chart(
        deck,
        height=820,
        width="stretch",
        on_select="rerun",
        selection_mode="single-object",
        key="railway-map",
    )

selected_objects = event.selection.objects.get("railway-stretches", [])
selected_object = selected_objects[0] if selected_objects else None
selected_properties = (
    selected_object.get("properties", selected_object) if selected_object else None
)
if not selected_properties and selected_search_entity and selected_search_entity["kind"] == "Section":
    selected_id = selected_search_entity["section_ids"][0]
    selected_feature = next(
        (
            feature
            for feature in railway_geometry["features"]
            if int(feature["properties"]["STRKAFS"]) == selected_id
        ),
        None,
    )
    selected_properties = selected_feature["properties"] if selected_feature else None
with details_panel:
    if selected_properties:
        render_contract_panel(selected_properties, contracts)
    elif selected_search_entity:
        render_search_entity_panel(selected_search_entity, contracts)
    else:
        render_contract_panel(None, contracts)
